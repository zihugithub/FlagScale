import os
import sys
import yaml
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(e)
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter


unit = {
    "name": "单元测试指令运行时间统计",
    "image": "镜像",
    'backend': [ "megatron", "flagscale" ],
    "subset": { 
        "megatron": ["data", "dist_checkpointing", "distributed", "export", "fusions", "inference",
               "models", "pipeline_parallel", "post_training", "ssm", "tensor_parallel", "transformer/moe",
               "transformer", "./", "", "" ],
        "flagscale": [ "runner",  "./", "", "" ]
    },
    "status": [ "offline", "online" ],
    "runtime": "case_runtime"
}

functional = {
    "name": "功能测试指令case运行时间统计",
    "image": "镜像",
    "type": [ "inference", "train", "hetero_train", "serve" ],
    "task": {
        "inference": [ "deepseek", "qwen3", "deepseek_flaggems", "qwen3_flaggems" ],
        "train": [ "aquila", "deepseek", "mixtral", "llava_onevision" ],
        "hetero_train": [ "aquila" ],
        "serve": [ "base" ]
    },
    "case_runtime": {
        "inference": {
            "deepseek": [ "tp2", "tp4" ],
            "qwen3": [ "tp2", "tp4" ],
            "deepseek_flaggems": [ "tp2" ],
            "qwen3_flaggems": [ "tp2" ]
        },
        "train": {
             "aquila": [ "tp2_pp2", "tp4_pp2" ],
             "deepseek": [ "tp2_pp2_ep2" ], 
             "mixtral": [ "tp2_pp1_ep2", "tp4_pp1_ep2" ],
             "llava_onevision": [ "tp2", "tp4" ]
        } ,
        "hetero_train": {
            "aquila": [ "tp2pp1_tp4pp1_tp2pp1", "tp2dp1pp1_tp2dp2pp1_tp1dp2pp1", "dp2dp4_shared_embedding" ]
        },
        "serve": {
            "base": [ "multiple_model" ]
        }
    }
}

class test_time_statistics(object):
    def __init__(self,oper_type):
        self.file_name_txt = f"{oper_type}_runtime.txt"
        self.file_name_xlsx = f"{oper_type}-test.time.xlsx"

    def xlsx_path(self): return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    def data_file(self):
        '''
        File Content Format:
        - tests/unit_runtime.txt
          eg: >>>: tests/scripts/unit_tests/test_subset.sh --backend megatron --subset data runtime: 2
        - tests/functional_runtime.txt
          eg: >>>: train_aquila_tp2_pp2 runtime: 2
        '''
        for root,dirs,files in os.walk(os.getcwd()):
            if self.file_name_txt in files: 
                self.file_name_txt = os.path.join(root,self.file_name_txt)

        assert os.path.isfile(self.file_name_txt), f"请先将测试用例运行时间存于文件：{self.xlsx_path()}/{self.file_name_txt}"
        return self.file_name_txt

    def get_image(self):
        '''
        NOTE: 镜像需要手动填写【没权限读取环境变量】
        DATE: 2025-06-06
        '''
        workflows = f"{os.path.dirname(self.xlsx_path())}/.github/workflows/all-tests.yml"
        assert os.path.isfile(workflows), f"请确保文件 {workflows} 的存在"
        with open(workflows,"r") as file:
            config = yaml.safe_load(file)
        if 'jobs' not in config:
            raise ValueError(f"Test node 'jobs' not found in configuration file.")
        else:
            jobs = config['jobs']

        if "set-env" not in jobs:
            raise ValueError(f"Test node 'jobs.set-env' not found in configuration file.")
        else:
            set_env = jobs["set-env"]

        if 'steps' not in set_env:
            raise ValueError(f"Test node 'jobs.set-env.steps' not found in configuration file.")
        else:
            steps = set_env["steps"]

        if 'run' not in steps[0]:
            raise ValueError(f"Test node 'jobs.set-env.steps.[run]' not found in configuration file.")
        else:
            run = steps[0]['run'].strip()

        if "ci_image" not in run:
            raise ValueError(f"Test node 'jobs.set-env.steps.[run].ci_image' not found in configuration file.")
        else:
            ci_image = run.split('"')[1].split("=")[1]

        return ci_image

    def create_unit_layout(self):
        wb = Workbook()
        ws = wb.active
        # 单元测试指令运行时间统计 Field Arrangement
        ws.merge_cells("A1:C1")
        ws['A1'] = unit["name"]
        # backend Field Arrangement
        ws.merge_cells("A2:A3")
        ws['A2'] = list(unit.keys())[2]
        # subset Field Arrangement
        ws.merge_cells("B2:B3")
        ws['B2'] = list(unit.keys())[3]
        # status Field Arrangement
        ws.merge_cells("C2:C3")
        ws['C2'] = list(unit.keys())[4]
        # 镜像 Field Arrangement
        ws['D1'] = unit["image"]
        # case_runtime Field Arrangement
        ws['D3'] = unit["runtime"]

        # Subset megatron Field Arrangement
        for megatron_index0, megatron_site in enumerate(range(4, len(unit["subset"]["megatron"]) + 4)):
            if megatron_index0 == 0:
                megatron_start_index = megatron_site
            elif megatron_index0 == len(range(4, len(unit["subset"]["megatron"]) + 4)) - 1:
                megatron_end_index = megatron_site
            for megatron_index1, megatron_field in enumerate(unit["subset"]["megatron"]):
                if megatron_index0 == megatron_index1:
                    ws["B%s"%(megatron_site)] = megatron_field
                    ws["D%s"%(megatron_site)] = 2  # eg
        ws["D%s"%(megatron_end_index)] = "none"    # eg

        # Subset Flagscale Field Arrangement
        for flagscale_index0, flagscale_site in enumerate(range(megatron_end_index + 1, megatron_end_index + len(unit["subset"]["flagscale"]) + 1)):
            if flagscale_index0 == 0:
                flagscale_start_index = flagscale_site
            flagscale_end_index = megatron_end_index + len(unit["subset"]["flagscale"])
            for flagscale_index1, flagscale_field in enumerate(unit["subset"]["flagscale"]):
                if flagscale_index0 == flagscale_index1:
                    ws["B%s"%(flagscale_site)] = flagscale_field
                    ws["D%s"%(flagscale_site)] = 2  # eg
        ws["D%s"%(flagscale_end_index)] = "none"    # eg
        ws["D2"] = "flagscale:cuda12.4.1-cudnn9.5.0-python3.12-torch2.6.0-time2505241715"

        # backend megatron Field Arrangement
        ws.merge_cells("A{}:A{}".format(megatron_start_index,megatron_end_index))
        ws["A%s"%(megatron_start_index)] = list(unit["subset"].keys())[0]

        # status megatron Field Arrangement
        ws.merge_cells("C{}:C{}".format(megatron_start_index,megatron_end_index-2))
        ws["C%s"%(megatron_end_index-1)] = unit["status"][0]
        ws["C%s"%(megatron_end_index)] = unit["status"][1]

        # backend Flagscale Field Arrangement
        ws.merge_cells("A{}:A{}".format(flagscale_start_index, flagscale_end_index))
        ws["A%s"%(flagscale_start_index)] = list(unit["subset"].keys())[1]

        # status flagscale Field Arrangement
        ws.merge_cells("C{}:C{}".format(flagscale_start_index,flagscale_end_index-2))
        ws["C%s"%(flagscale_end_index-1)] = unit["status"][0]
        ws["C%s"%(flagscale_end_index)] = unit["status"][1]
        # eg
        wb.save(f"{self.xlsx_path()}/{self.file_name_xlsx}")

    def create_functional_layout(self):
        wb = Workbook()
        ws = wb.active
        # 功能测试指令case运行时间统计 Field Arrangement
        ws.merge_cells("A1:B1")
        ws['A1'] = unit["name"]
        # type Field Arrangement
        ws['A2'] = list(functional.keys())[2]
        # task Field Arrangement
        ws['B2'] = list(functional.keys())[3]
        # case_runtime Field Arrangement
        ws['C2'] = list(functional.keys())[4]
        # 镜像 Field Arrangement
        ws['C1'] = functional["image"]

        # task inference Field Arrangement
        for inference_index0, inference_site in enumerate(range(3, len(functional["task"]["inference"]) + 3)):
            if inference_index0 == 0:
                inference_start_index = inference_site
            elif inference_index0 == len(range(3, len(functional["type"]) + 3 )) - 1:
                inference_end_index = inference_site
            for inference_index1, inference_field in enumerate(functional["task"]["inference"]):
                if inference_index0 == inference_index1:
                    ws["B%s"%(inference_site)] = inference_field

        # task train Field Arrangement
        for train_index0, train_site in enumerate(range(inference_end_index + 1, inference_end_index + len(functional["task"]["train"]) + 1)):
            if train_index0 == 0:
                train_start_index = train_site

            train_end_index = inference_end_index + len(functional["task"]["train"])
            for train_index1, train_field in enumerate(functional["task"]["train"]):
                if train_index0 == train_index1:
                    ws["B%s"%(train_site)] = train_field

        # task hetero_train Field Arrangement
        for hetero_train_index0, hetero_train_site in enumerate(range(train_end_index + 1, train_end_index + len(functional["task"]["hetero_train"]) + 1)):
            if hetero_train_index0 == 0:
                hetero_train_start_index = hetero_train_site

            hetero_train_end_index = train_end_index + len(functional["task"]["hetero_train"])
            for hetero_train_index1, hetero_train_field in enumerate(functional["task"]["hetero_train"]):
                if hetero_train_index0 == hetero_train_index1:
                    ws["B%s"%(hetero_train_site)] = hetero_train_field

        # task serve Field Arrangement
        for serve_index0, serve_site in enumerate(range(hetero_train_end_index + 1, hetero_train_end_index + len(functional["task"]["serve"]) + 1)):
            if serve_index0 == 0:
                serve_start_index = serve_site

            serve_end_index = hetero_train_end_index + len(functional["task"]["serve"])
            for serve_index1, serve_field in enumerate(functional["task"]["serve"]):
                if serve_index0 == serve_index1:
                    ws["B%s"%(serve_site)] = serve_field

        # type inference Field Arrangement
        ws.merge_cells("A{}:A{}".format(inference_start_index,inference_end_index))
        ws["A{}".format(inference_start_index)] = functional["type"][0]

        # type train Field Arrangement
        ws.merge_cells("A{}:A{}".format(train_start_index,train_end_index))
        ws["A{}".format(train_start_index)] = functional["type"][1]

        # type hetero_train Field Arrangement
        ws.merge_cells("A{}:A{}".format(hetero_train_start_index,hetero_train_end_index))
        ws["A{}".format(hetero_train_start_index)] = functional["type"][2]

        # type serve Field Arrangement
        ws.merge_cells("A{}:A{}".format(serve_start_index,serve_end_index))
        ws["A{}".format(serve_start_index)] = functional["type"][3]

        # case_runtime Field Arrangement
        li_case_runtime = []
        for type in list(functional["case_runtime"]):
            for task in list(functional["case_runtime"][type]):
                if task in list(functional["case_runtime"][type]):
                    case_runtime = "\n".join(f"{case_runtime}:" for case_runtime in list(functional["case_runtime"][type][task]))
                    li_case_runtime.append(case_runtime)

        for case_runtime_index0,case_runtime_site in enumerate(range(3,sum(len(lst) for lst in functional["task"].values()) + 3)):
            for case_runtime_index1,case_runtime_field in enumerate(li_case_runtime):
                if case_runtime_index0 == case_runtime_index1:
                    ws["C%s"%(case_runtime_site)] = case_runtime_field

        wb.save(f"{self.xlsx_path()}/{self.file_name_xlsx}")

    def get_line_range(self):
        wb = load_workbook(f"{self.xlsx_path()}/{self.file_name_xlsx}")
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                if "case_runtime" in str(cell.value):
                    start_case_runtime_site = cell.row + 1
                    break  # 避免同一行多次匹配

        end_case_runtime_site = 0
        for sheet in wb.worksheets:
            end_case_runtime_site += sheet.max_row

        return start_case_runtime_site, end_case_runtime_site

    def unit_add_data(self):
        wb = load_workbook(f"{self.xlsx_path()}/{self.file_name_xlsx}")
        ws = wb.active
        max_columns = get_column_letter(ws.max_column) # 整个工作表中所有数据的最大列号
        for action_column in range(ord(max_columns) + 1, ord(max_columns) + 2):  # 遍历大写字母
            action_column = chr(action_column)

        # 遍历所有合并区域
        for merged_cells in ws.merged_cells:
            min_sheet0 = merged_cells.coord.split(":")[0]
            max_sheet1 = merged_cells.coord.split(":")[1]
            letter0 = "".join([char for char in min_sheet0 if char.isalpha()])
            number0 = "".join([char for char in min_sheet0 if char.isdigit()])
            letter1 = "".join([char for char in max_sheet1 if char.isalpha()]) # A
            number1 = "".join([char for char in max_sheet1 if char.isdigit()]) # 19
            if letter0 == letter1 and letter1 == "A" and int(number1) >= 4:
                backend = ws[min_sheet0].value
                for sub_set_index in range(int(number0), int(number1) + 1):
                    subset = ws[f"B{sub_set_index}"].value    
                    com_key_xlsx = "_".join([str(backend), str(subset)])

        with open(self.data_file(),"r") as f:
            lines = f.readlines()
            for line in lines:
                if "runtime" in line:
                    com_key_txt = line.split(r' ')[3] + "_" + line.split(r' ')[5]
                    runtime = line.split(r' ')[7].strip()

                    for merged_cells in ws.merged_cells:
                        min_sheet0 = merged_cells.coord.split(":")[0]
                        max_sheet1 = merged_cells.coord.split(":")[1]
                        letter0 = "".join([char for char in min_sheet0 if char.isalpha()])
                        number0 = "".join([char for char in min_sheet0 if char.isdigit()])
                        letter1 = "".join([char for char in max_sheet1 if char.isalpha()]) # A
                        number1 = "".join([char for char in max_sheet1 if char.isdigit()]) # 19
                        if letter0 == letter1 and letter1 == "A" and int(number1) >= 4:
                            backend = ws[min_sheet0].value
                            for sub_set_index in range(int(number0), int(number1) + 1):
                                subset = ws[f"B{sub_set_index}"].value
                                com_key_xlsx = "_".join([str(backend), str(subset)])
                                if com_key_txt == com_key_xlsx:
                                    ws[f"{action_column}{sub_set_index}"] = runtime

        start_case_runtime_site = self.get_line_range()[0]
        end_case_runtime_site = self.get_line_range()[1]
        for activate_site in range(start_case_runtime_site, end_case_runtime_site + 1):
            if ws[f"{action_column}{activate_site}"].value == None :
                ws[f"{action_column}{activate_site}"] = "none"

        wb.save(f"{self.xlsx_path()}/{self.file_name_xlsx}")


    def functional_add_data(self):
        wb = load_workbook(f"{self.xlsx_path()}/{self.file_name_xlsx}")
        ws = wb.active
        max_columns = get_column_letter(ws.max_column) # 整个工作表中所有数据的最大列号
        action_column = chr(ord(max_columns) + 1) # 新增列号，如：D

        com_key_xlsx_list = {} # 从 xlsx 中获取行号(key) 及 case_runtime
        for merged_cells in ws.merged_cells:
            if len(merged_cells.coord.split(":")) == 1:  # ['A12'] 
                max_sheet0 = merged_cells.coord.split(":")[0]  # A12
                letter0 = "".join([char for char in max_sheet0 if char.isalpha()])  # A
                number0 = "".join([char for char in max_sheet0 if char.isdigit()])  # 12
                if letter0 == "A" and int(number0) >= 3:
                    _type = ws[max_sheet0].value      # server
                    _task = ws[f"B{number0}"].value   # base            
                    _case_runtime = ws[f"C{number0}"].value.split("\n")  # ['multiple_model:']
                    for i in range(0, len(_case_runtime)):
                        com_key_xlsx = "_".join([str(_type), str(_task), str(_case_runtime[i])]) # serve_base_multiple_model
                        com_name = "_".join([str(_type), str(_task)])
                        com_key_xlsx_list.setdefault(number0,[]).append(f"{com_key_xlsx} {_case_runtime[i]}")

            elif len(merged_cells.coord.split(":")) == 2: # ['A4', 'A19']
                min_sheet0 = merged_cells.coord.split(":")[0]
                max_sheet1 = merged_cells.coord.split(":")[1]
                letter0 = "".join([char for char in min_sheet0 if char.isalpha()])
                number0 = "".join([char for char in min_sheet0 if char.isdigit()])
                letter1 = "".join([char for char in max_sheet1 if char.isalpha()]) # A
                number1 = "".join([char for char in max_sheet1 if char.isdigit()]) # 19
                if letter0 == letter1 and letter1 == "A" and int(number1) >= 4:
                    _type = ws[min_sheet0].value
                    for sub_set_index in range(int(number0), int(number1) + 1):
                        _task = ws[f"B{sub_set_index}"].value
                        _case_runtime = ws[f"C{sub_set_index}"].value.split("\n")
                        for i in range(0, len(_case_runtime)):
                            com_key_xlsx = "_".join([str(_type), str(_task), str(_case_runtime[i])])  # serve_base_multiple_model
                            com_name = "_".join([str(_type), str(_task)])
                            com_key_xlsx_list.setdefault(sub_set_index,[]).append(f"{com_key_xlsx} {_case_runtime[i]}")

        with open(self.data_file(),"r") as f:
            lines = f.readlines()
            cases_runtime = [] # ['train_aquila_tp2_pp2: 1', 'train_aquila_tp4_pp2: 2']
            for line in lines:
                if "runtime" in line:
                    runtime = "".join(f"{line.split(r' ')[1].strip()}: {line.split(r' ')[3].strip()}") 
                    cases_runtime.append(runtime)

        site_and_field = {}
        for key_site,value_field_list in com_key_xlsx_list.items():
            for field in value_field_list: # serve_base_multiple_model: serve_base multiple_model:
                field0 = field.split(r' ')[0].strip()
                field1 = field.split(r' ')[1].strip()

                for case_runtime in cases_runtime: # 'serve_base_multiple_model: 2'
                    case_runtime0 = case_runtime.split(r' ')[0].strip()
                    case_runtime1 = case_runtime.split(r' ')[1].strip()
                    if field0 == case_runtime0:
                        site_and_field.setdefault(key_site,[]).append(f"{field1} {case_runtime1}")

        for site, field_cases in site_and_field.items():
            str_runtime = ''
            for field_case in field_cases:
                str_runtime += f"{field_case}\n"
            ws[f"{action_column}{site}"] = str_runtime

        start_case_runtime_site = self.get_line_range()[0]
        end_case_runtime_site = self.get_line_range()[1]
        for activate_site in range(start_case_runtime_site, end_case_runtime_site + 1):
            if ws[f"{action_column}{activate_site}"].value == None :
                ws[f"{action_column}{activate_site}"] = "none"

        wb.save(f"{self.xlsx_path()}/{self.file_name_xlsx}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: test.time_statistics.py <test_type>")
        sys.exit(1)
    test_type = sys.argv[1]

    # test_type = "functional"
    # test_type = "unit"

    case = test_time_statistics(test_type)

    if os.path.isfile(case.file_name_xlsx):
        if test_type == "unit":
            case.unit_add_data()
        elif test_type == "functional":
            case.functional_add_data()
    else:
        if test_type == "unit":
            case.create_unit_layout()
            case.get_line_range()
            case.unit_add_data()
        elif test_type == "functional":
            case.create_functional_layout()
            case.get_line_range()
            case.functional_add_data()
